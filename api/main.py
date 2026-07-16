"""
main.py
-------
The ENTIRE backend in one file. No other Python files needed
(except the bundled AI_Financial_Report_Template.xlsx, which must
sit next to this file at deploy time).

UPGRADE NOTES (read this before touching anything):
  - AI_Financial_Report_Template.xlsx is the single source of truth for
    financial statement structure. It is NEVER regenerated, redesigned,
    or given new line items. It is only ever copied, and only its
    VALUES are ever changed.
  - The old "manual entry" flow (arbitrary particulars, freeform amounts)
    has been replaced by an AI Accounting Assistant: the user types a
    plain-English transaction, Gemini classifies it against the fixed
    line items in the template, and only those exact line items are
    ever updated.
  - A lightweight "period" concept is kept (Fiscal Year + Quarter),
    because the dashboard's QoQ / YoY trend reporting depends on it.
    Each period gets its own copy of the three statement sheets,
    cloned from the master template the first time that period is used.
  - Everything is still stored as a single .xlsx file in a GitHub repo
    via the Contents API - no database.

Environment variables needed (set these in Vercel Project Settings):
  GITHUB_TOKEN    - GitHub Personal Access Token (Contents: Read & Write)
  GITHUB_REPO     - "your-username/your-repo"
  GITHUB_BRANCH   - usually "main"
  EXCEL_PATH      - where the live .xlsx lives in the repo, e.g. "data/finance_data.xlsx"
  ACCESS_CODE     - the shared password used to log in
  GEMINI_API_KEY  - Google Gemini API key (from .env / platform secrets - never hardcoded)
  GEMINI_MODEL    - optional, defaults to "gemini-2.0-flash"
"""

import os
import io
import re
import json
import base64
from datetime import datetime, timezone
from typing import Dict, List, Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import Series, SeriesLabel
from openpyxl.chart.data_source import NumDataSource, NumRef, StrRef, AxDataSource

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel


# ============================================================================
# SETTINGS
# ============================================================================
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "")
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "data/finance_data.xlsx")
ACCESS_CODE = os.environ.get("ACCESS_CODE", "finance2026")

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
# Pinned model versions (gemini-2.0-flash, gemini-2.5-flash-lite, etc.) keep
# getting closed to new users as Google rotates its lineup. The "-latest"
# alias always points at whatever Google currently recommends for that tier,
# so it's the safer default. Override via env var if you want a specific
# pinned version once you've confirmed it's available to your account.
GEMINI_MODEL = os.environ.get("GEMINI_MODEL") or "gemini-flash-lite-latest"
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

# The master template ships alongside this deployment. Where exactly it lands
# depends on your platform's bundling rules (e.g. Vercel's `includeFiles` is
# relative to the PROJECT ROOT, not to this file's own folder), so we check
# every likely location instead of assuming one.
TEMPLATE_FILENAME = "AI_Financial_Report_Template.xlsx"
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_CANDIDATE_TEMPLATE_PATHS = [
    os.path.join(_THIS_DIR, TEMPLATE_FILENAME),                    # next to main.py (e.g. api/)
    os.path.join(_THIS_DIR, "..", TEMPLATE_FILENAME),              # project root, one level up from api/
    os.path.join(os.getcwd(), TEMPLATE_FILENAME),                  # wherever the process's cwd is
    os.path.join("/var/task", TEMPLATE_FILENAME),                  # Vercel's function root
    os.path.join("/var/task", "api", TEMPLATE_FILENAME),           # Vercel's function root, api/ subfolder
]


def _resolve_template_path():
    for path in _CANDIDATE_TEMPLATE_PATHS:
        if os.path.exists(path):
            return os.path.abspath(path)
    return None


TEMPLATE_PATH = _resolve_template_path()

GITHUB_API = "https://api.github.com"
GH_HEADERS = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json",
}

QUARTER_ORDER = ["Q1", "Q2", "Q3", "Q4"]

# Sheet-name prefixes for the three statements, per period.
STATEMENT_PREFIX = {"Profit & Loss": "PL", "Balance Sheet": "BS", "Cash Flow": "CF"}
STATEMENT_LABELS = list(STATEMENT_PREFIX.keys())  # canonical order

LOG_SHEET = "Transaction Log"
# Extra Fiscal Year / Quarter columns are prepended so history stays filterable
# by period; the remaining seven columns match the spec exactly.
LOG_HEADERS = ["Fiscal Year", "Quarter", "Date", "Original Transaction",
               "Statement", "Line Item", "Amount", "Operation", "Status"]

# P&L rows that are computed by formula in the template - never write to these
# directly, they are derived in Python from the editable rows around them.
PL_COMPUTED_ROWS = {
    "Total Income", "Total Expenses", "Profit Before Exceptional Items & Tax",
    "Profit Before Tax", "Profit After Tax",
}
BS_COMPUTED_ROWS = {"TOTAL ASSETS", "TOTAL EQUITY & LIABILITIES"}
CF_COMPUTED_ROWS = {"Closing Cash"}


# ============================================================================
# GITHUB READ / WRITE (the "database")
# ============================================================================
def _contents_url(path):
    return f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{path}"


def _get_file_meta():
    """Returns (sha, raw_bytes). raw_bytes is None if the file doesn't exist yet."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        raise RuntimeError("GITHUB_TOKEN / GITHUB_REPO are not set on the server.")
    resp = requests.get(_contents_url(EXCEL_PATH), headers=GH_HEADERS, params={"ref": GITHUB_BRANCH}, timeout=30)
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    data = resp.json()
    return data["sha"], base64.b64decode(data["content"])


def _put_file(raw_bytes, commit_message):
    sha, _ = _get_file_meta()
    payload = {
        "message": commit_message,
        "content": base64.b64encode(raw_bytes).decode(),
        "branch": GITHUB_BRANCH,
    }
    if sha:
        payload["sha"] = sha
    resp = requests.put(_contents_url(EXCEL_PATH), headers=GH_HEADERS, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()


def get_download_url():
    return f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{EXCEL_PATH}"


# ============================================================================
# EXCEL SERVICE
# Everything that touches the master template and the live workbook.
# Line items are always located by ROW LABEL, never by hardcoded coordinates,
# so rows can move in the template without breaking the code.
# ============================================================================
class ExcelService:

    @staticmethod
    def load_master_template() -> Workbook:
        if not TEMPLATE_PATH:
            tried = ", ".join(_CANDIDATE_TEMPLATE_PATHS)
            raise RuntimeError(
                f"{TEMPLATE_FILENAME} could not be found in any of the expected locations. "
                f"Tried: {tried}. cwd={os.getcwd()!r}. "
                "Check your deployment's file-bundling config (e.g. vercel.json includeFiles) "
                "matches where this file actually lands."
            )
        return load_workbook(TEMPLATE_PATH, data_only=False)

    @classmethod
    def get_line_items(cls) -> Dict[str, List[str]]:
        """
        {"Profit & Loss": [...editable row labels...], "Balance Sheet": [...], "Cash Flow": [...]}
        Skips section headers (amount cell blank) and computed/formula rows -
        those are exactly the rows Gemini must never be offered as a target.
        """
        wb = cls.load_master_template()
        out = {}
        for sheet in STATEMENT_LABELS:
            ws = wb[sheet]
            items = []
            for row in ws.iter_rows(min_row=2, max_col=2):
                label_cell, amount_cell = row[0], row[1]
                if label_cell.value is None:
                    continue
                if amount_cell.value is None:
                    continue  # section header, e.g. "ASSETS"
                if isinstance(amount_cell.value, str) and amount_cell.value.startswith("="):
                    continue  # computed/total row
                items.append(label_cell.value)
            out[sheet] = items
        return out

    @classmethod
    def load_live_workbook(cls) -> Workbook:
        _, raw = _get_file_meta()
        if raw is None:
            wb = Workbook()
            ws = wb.active
            ws.title = LOG_SHEET
            ws.append(LOG_HEADERS)
            return wb
        wb = load_workbook(io.BytesIO(raw), data_only=False)
        if LOG_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(LOG_SHEET)
            ws.append(LOG_HEADERS)
        return wb

    @staticmethod
    def period_sheet_name(statement: str, fiscal_year: int, quarter: str) -> str:
        return f"{STATEMENT_PREFIX[statement]}_{fiscal_year}_{quarter}"

    @staticmethod
    def previous_period(fiscal_year: int, quarter: str):
        """The chronologically preceding (fiscal_year, quarter), wrapping Q1 -> prior year's Q4."""
        idx = QUARTER_ORDER.index(quarter)
        if idx == 0:
            return fiscal_year - 1, QUARTER_ORDER[-1]
        return fiscal_year, QUARTER_ORDER[idx - 1]

    @classmethod
    def ensure_period(cls, wb: Workbook, fiscal_year: int, quarter: str):
        """Clone the three statement sheets from the master template into this
        period's sheets, the first time this period is touched. Values start
        at 0 (or blank for header rows); only VALUES are ever generated fresh -
        never new line items.

        Balance Sheet and Cash Flow are POINT-IN-TIME statements, not flows -
        a company's assets/liabilities/equity and cash balance don't reset to
        zero every quarter. So immediately after cloning, if a prior period
        exists, we carry its Balance Sheet line items and Cash Flow's Opening
        Cash forward into the new period. Profit & Loss is correctly left at
        zero - it's a flow statement that legitimately starts fresh each
        quarter."""
        template_wb = cls.load_master_template()
        created_any = False
        for statement in STATEMENT_LABELS:
            name = cls.period_sheet_name(statement, fiscal_year, quarter)
            if name in wb.sheetnames:
                continue
            created_any = True
            src: Worksheet = template_wb[statement]
            dst = wb.create_sheet(name)
            for row in src.iter_rows():
                for cell in row:
                    dst.cell(row=cell.row, column=cell.column, value=cell.value)
            for col_letter, dim in src.column_dimensions.items():
                dst.column_dimensions[col_letter].width = dim.width

        if created_any:
            cls._carry_forward_balances(wb, fiscal_year, quarter)

    @classmethod
    def _carry_forward_balances(cls, wb: Workbook, fiscal_year: int, quarter: str):
        """Seed a freshly-created period's Balance Sheet and Opening Cash from
        the immediately preceding period's closing figures, if that period
        exists. If there's no prior period (e.g. this is the very first
        quarter ever entered), the template's zero defaults are left as-is."""
        prev_fy, prev_q = cls.previous_period(fiscal_year, quarter)
        prev_bs_name = cls.period_sheet_name("Balance Sheet", prev_fy, prev_q)
        if prev_bs_name not in wb.sheetnames:
            return  # no prior period to carry forward from

        # Balance Sheet: carry every editable line item forward as-is.
        prev_bs_values = cls.get_period_values(wb, "Balance Sheet", prev_fy, prev_q)
        new_bs = wb[cls.period_sheet_name("Balance Sheet", fiscal_year, quarter)]
        for row in new_bs.iter_rows(min_row=2, max_col=2):
            label, cell = row[0].value, row[1]
            if label in prev_bs_values:
                cell.value = prev_bs_values[label]

        # Cash Flow: only Opening Cash carries forward (= prior quarter's
        # computed Closing Cash). Net Cash from Operating/Investing/Financing
        # are genuine quarterly flows and correctly stay at zero.
        prev_cf_values = cls.get_period_values(wb, "Cash Flow", prev_fy, prev_q)
        prev_closing_cash = AccountingService.cf_totals(prev_cf_values).get("Closing Cash", 0)
        new_cf = wb[cls.period_sheet_name("Cash Flow", fiscal_year, quarter)]
        for row in new_cf.iter_rows(min_row=2, max_col=2):
            if row[0].value == "Opening Cash":
                row[1].value = prev_closing_cash
                break

    @classmethod
    def find_cell(cls, wb: Workbook, statement: str, fiscal_year: int, quarter: str, line_item: str):
        name = cls.period_sheet_name(statement, fiscal_year, quarter)
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[0].value == line_item:
                return ws, row[1]
        return None, None

    @classmethod
    def apply_entry(cls, wb: Workbook, fiscal_year: int, quarter: str,
                     statement: str, line_item: str, amount: float, operation: str) -> float:
        cls.ensure_period(wb, fiscal_year, quarter)
        ws, cell = cls.find_cell(wb, statement, fiscal_year, quarter, line_item)
        if cell is None:
            raise ValueError(f"Line item '{line_item}' not found in '{statement}'.")
        current = cell.value or 0
        if isinstance(current, str):
            raise ValueError(f"'{line_item}' is a computed row and cannot be edited directly.")
        if operation == "add":
            new_val = round(float(current) + float(amount), 2)
        elif operation == "subtract":
            new_val = round(float(current) - float(amount), 2)
        else:
            raise ValueError(f"Unknown operation '{operation}'.")
        cell.value = new_val
        return new_val

    @classmethod
    def get_period_values(cls, wb: Workbook, statement: str, fiscal_year: int, quarter: str) -> Dict[str, float]:
        name = cls.period_sheet_name(statement, fiscal_year, quarter)
        if name not in wb.sheetnames:
            return {}
        ws = wb[name]
        out = {}
        for row in ws.iter_rows(min_row=2, max_col=2):
            label, amount = row[0].value, row[1].value
            if label is None or amount is None:
                continue
            if isinstance(amount, str):
                continue  # formula/computed row, derived separately
            out[label] = float(amount)
        return out

    @classmethod
    def list_periods(cls, wb: Workbook):
        """[(fiscal_year, quarter), ...] for every period that has at least one statement sheet."""
        seen = set()
        pattern = re.compile(r"^(?:PL|BS|CF)_(\d{4})_(Q[1-4])$")
        for name in wb.sheetnames:
            m = pattern.match(name)
            if m:
                seen.add((int(m.group(1)), m.group(2)))
        return sorted(seen, key=lambda t: (t[0], QUARTER_ORDER.index(t[1])))

    @classmethod
    def append_log(cls, wb: Workbook, fiscal_year: int, quarter: str, original_txn: str,
                    statement: str, line_item: str, amount: float, operation: str, status: str):
        ws = wb[LOG_SHEET]
        if ws.cell(row=1, column=1).value != LOG_HEADERS[0]:
            for i, h in enumerate(LOG_HEADERS, start=1):
                ws.cell(row=1, column=i, value=h)
        now = datetime.now(timezone.utc).isoformat()
        ws.append([fiscal_year, quarter, now, original_txn, statement, line_item, amount, operation, status])
        return ws.max_row

    @classmethod
    def get_log_rows(cls, wb: Workbook, fiscal_year: Optional[int] = None, quarter: Optional[str] = None):
        ws = wb[LOG_SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        out = [dict(zip(LOG_HEADERS, r)) for r in rows if r[0] is not None]
        if fiscal_year is not None:
            out = [r for r in out if int(r["Fiscal Year"]) == fiscal_year]
        if quarter is not None:
            out = [r for r in out if r["Quarter"] == quarter]
        return out

    @classmethod
    def undo_last(cls, wb: Workbook, fiscal_year: int, quarter: str):
        ws = wb[LOG_SHEET]
        target_row = None
        for r in range(ws.max_row, 1, -1):
            if (ws.cell(row=r, column=1).value == fiscal_year
                    and ws.cell(row=r, column=2).value == quarter
                    and ws.cell(row=r, column=9).value == "Applied"):
                target_row = r
                break
        if target_row is None:
            raise ValueError("No transaction to undo for this period.")

        statement = ws.cell(row=target_row, column=5).value
        line_item = ws.cell(row=target_row, column=6).value
        amount = float(ws.cell(row=target_row, column=7).value)
        operation = ws.cell(row=target_row, column=8).value
        original_txn = ws.cell(row=target_row, column=4).value

        reverse_op = "subtract" if operation == "add" else "add"
        cls.apply_entry(wb, fiscal_year, quarter, statement, line_item, amount, reverse_op)
        ws.cell(row=target_row, column=9, value="Reversed")
        cls.append_log(wb, fiscal_year, quarter, f"UNDO: {original_txn}", statement,
                        line_item, amount, reverse_op, "Applied (undo)")
        return {"statement": statement, "line_item": line_item, "amount": amount, "reversed_operation": operation}

    @classmethod
    def backfill_carry_forward(cls, wb: Workbook):
        """ONE-TIME REPAIR for periods created before carry-forward existed.
        Any period after the very first currently holds only the transactions
        entered directly against it, sitting on a zero Balance Sheet baseline
        and zero Opening Cash. This walks periods chronologically and ADDS
        the correct prior period's closing balances on top of whatever was
        already entered - it never overwrites entered transaction deltas.
        Profit & Loss (a flow statement) is never touched.

        Safe to run once. Do NOT run twice - a second run would add the
        carried-forward balances again on top of the now-corrected values."""
        periods = cls.list_periods(wb)
        fixed = []
        for i in range(1, len(periods)):
            fy, q = periods[i]
            prev_fy, prev_q = periods[i - 1]

            prev_bs_values = cls.get_period_values(wb, "Balance Sheet", prev_fy, prev_q)
            bs_ws = wb[cls.period_sheet_name("Balance Sheet", fy, q)]
            for row in bs_ws.iter_rows(min_row=2, max_col=2):
                label, cell = row[0].value, row[1]
                if label in prev_bs_values:
                    current = cell.value or 0
                    if isinstance(current, str):
                        continue  # formula/computed row, skip
                    cell.value = round(float(current) + prev_bs_values[label], 2)

            prev_cf_values = cls.get_period_values(wb, "Cash Flow", prev_fy, prev_q)
            prev_closing_cash = AccountingService.cf_totals(prev_cf_values).get("Closing Cash", 0)
            cf_ws = wb[cls.period_sheet_name("Cash Flow", fy, q)]
            for row in cf_ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == "Opening Cash":
                    row[1].value = prev_closing_cash
                    break

            fixed.append({"fiscal_year": fy, "quarter": q})
        return fixed

    @classmethod
    def save(cls, wb: Workbook, message: str):
        buf = io.BytesIO()
        wb.save(buf)
        return _put_file(buf.getvalue(), message)


# ============================================================================
# ACCOUNTING SERVICE
# The accounting-classification knowledge handed to Gemini, and the pure
# Python formulas that mirror the template's totals (openpyxl doesn't
# evaluate formulas, so totals for the API/dashboard are computed here from
# the same named rows the template's own formulas reference).
# ============================================================================
class AccountingService:

    CLASSIFICATION_HINTS = """
Common classification examples (use the exact line item names given to you,
never these example names verbatim unless they also appear in your list):
  Salary / wages paid              -> Employee Benefits Expense
  Laptop / computer / equipment    -> Property, Plant & Equipment (Balance Sheet)
  Raw material / inventory buy     -> Inventories (Balance Sheet) AND Cost of Materials Consumed (P&L) if consumed
  Finished goods sold              -> Revenue From Operations (P&L)
  Interest received                -> Other Income (P&L)
  Interest paid / loan interest    -> Finance Costs (P&L)
  Loan taken                       -> Borrowings (Balance Sheet)
  Supplier credit / bill payable   -> Trade Payables (Balance Sheet)
  Tax paid                         -> Current Tax (P&L)
  Customer payment collected       -> Cash & Cash Equivalents (Balance Sheet) and/or Trade Receivables
  Rent, utilities, misc overhead   -> Other Expenses (P&L)
  Depreciation                     -> Depreciation & Amortisation Expenses (P&L)
"""

    @staticmethod
    def pl_totals(values: Dict[str, float]) -> Dict[str, float]:
        total_income = values.get("Revenue From Operations", 0) + values.get("Other Income", 0)
        expense_rows = ["Cost of Materials Consumed", "Purchases of Stock-in-Trade", "Changes in Inventories",
                         "Employee Benefits Expense", "Finance Costs",
                         "Depreciation & Amortisation Expenses", "Other Expenses"]
        total_expenses = sum(values.get(r, 0) for r in expense_rows)
        pbeit = total_income - total_expenses
        pbt = pbeit - values.get("Exceptional Items", 0)
        pat = pbt - values.get("Current Tax", 0) - values.get("Deferred Tax", 0)
        margin = (pat / total_income * 100) if total_income else 0
        return {
            "Total Income": total_income, "Total Expenses": total_expenses,
            "Profit Before Exceptional Items & Tax": pbeit, "Profit Before Tax": pbt,
            "Profit After Tax": pat, "Net Margin %": margin,
        }

    @staticmethod
    def bs_totals(values: Dict[str, float]) -> Dict[str, float]:
        asset_rows = ["Property, Plant & Equipment", "Right of Use Assets", "Investment Property",
                      "Other Intangible Assets", "Inventories", "Trade Receivables",
                      "Cash & Cash Equivalents", "Other Current Assets"]
        eq_liab_rows = ["Equity Share Capital", "Other Equity", "Borrowings",
                        "Trade Payables", "Other Current Liabilities"]
        return {
            "TOTAL ASSETS": sum(values.get(r, 0) for r in asset_rows),
            "TOTAL EQUITY & LIABILITIES": sum(values.get(r, 0) for r in eq_liab_rows),
        }

    @staticmethod
    def cf_totals(values: Dict[str, float]) -> Dict[str, float]:
        closing = (values.get("Net Cash from Operating", 0) + values.get("Net Cash from Investing", 0)
                   + values.get("Net Cash from Financing", 0) + values.get("Opening Cash", 0))
        return {"Closing Cash": closing}


# ============================================================================
# GEMINI SERVICE
# Sends the user's plain-English transaction + the fixed catalog of line
# items to Gemini and gets back structured JSON. Gemini may choose ONLY from
# the given line items, and must calculate totals (qty x price) itself.
# ============================================================================
class GeminiService:

    SYSTEM_TEMPLATE = """You are an expert Chartered Accountant acting as an AI bookkeeping assistant for an Indian company.
A user will describe one or more financial transactions in plain English. Your job:

1. Understand the transaction(s).
2. Calculate any totals yourself (e.g. "5 laptops @ 10000" = 50000; sum multiple items in one transaction).
3. Determine correct accounting treatment using standard Ind AS conventions.
4. Choose the correct SHEET and LINE ITEM only from the exact catalog given below - never invent a new line item or sheet name.
5. Decide the operation: "add" to increase a balance, "subtract" to decrease it.
6. If you are not confident which line item applies, do NOT guess: set "needs_confirmation": true,
   leave "entries" as an empty list, and explain what you need in "message".

Return ONLY a raw JSON object, nothing else - no markdown fences, no commentary, no preamble. Schema:
{
  "entries": [
    {"sheet": "<exact sheet name>", "line_item": "<exact line item from the catalog>", "amount": <positive number>, "operation": "add" or "subtract"}
  ],
  "needs_confirmation": false,
  "message": "<short optional note, empty string if none>"
}
"""

    @classmethod
    def build_prompt(cls, user_text: str, line_items_by_sheet: Dict[str, List[str]]) -> str:
        catalog = json.dumps(line_items_by_sheet, indent=2)
        return (
            f"{cls.SYSTEM_TEMPLATE}\n"
            f"{AccountingService.CLASSIFICATION_HINTS}\n"
            f"Catalog of sheets and their ONLY valid line items:\n{catalog}\n\n"
            f"User transaction:\n\"{user_text}\"\n\n"
            f"Return ONLY the JSON object described above."
        )

    @classmethod
    def parse_transaction(cls, user_text: str, line_items_by_sheet: Dict[str, List[str]]) -> dict:
        if not GEMINI_API_KEY:
            raise RuntimeError("GEMINI_API_KEY is not set on the server (add it to .env / platform secrets).")
        prompt = cls.build_prompt(user_text, line_items_by_sheet)
        resp = requests.post(
            GEMINI_URL,
            params={"key": GEMINI_API_KEY},
            headers={"x-goog-api-key": GEMINI_API_KEY, "Content-Type": "application/json"},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0, "response_mime_type": "application/json"},
            },
            timeout=30,
        )
        if resp.status_code == 404:
            raise RuntimeError(
                f"Gemini returned 404 for model '{GEMINI_MODEL}' on generateContent, even though this model "
                "appears in your key's ListModels response. This usually means the key needs to be sent as an "
                "'x-goog-api-key' header rather than a '?key=' query param (already tried both here), or the "
                "key is restricted to a different set of APIs/models than the ones ListModels shows. "
                f"Raw response: {resp.text[:500]}"
            )
        resp.raise_for_status()
        data = resp.json()
        try:
            text = data["candidates"][0]["content"]["parts"][0]["text"]
        except (KeyError, IndexError):
            raise RuntimeError("Gemini returned an unexpected response shape.")
        text = text.strip()
        text = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            raise RuntimeError("Gemini did not return valid JSON.")


# ============================================================================
# VALIDATION SERVICE
# Never trusts Gemini's output blindly. Rejects anything that doesn't match
# the master template's real sheets / line items, or has a bad amount/operation.
# ============================================================================
class ValidationService:
    ALLOWED_OPS = {"add", "subtract"}

    @classmethod
    def validate(cls, parsed: dict, line_items_by_sheet: Dict[str, List[str]]):
        """Returns (valid_entries, needs_confirmation, message)."""
        if not isinstance(parsed, dict):
            return [], False, "Gemini's response could not be understood."
        if parsed.get("needs_confirmation"):
            return [], True, parsed.get("message") or "I'm not confident how to classify this - could you clarify?"

        entries = parsed.get("entries") or []
        if not entries:
            return [], False, parsed.get("message") or "No valid entries were extracted from that transaction."

        valid, errors = [], []
        for e in entries:
            sheet = e.get("sheet")
            item = e.get("line_item")
            amount = e.get("amount")
            op = e.get("operation")
            if sheet not in line_items_by_sheet:
                errors.append(f"Unknown sheet '{sheet}'."); continue
            if item not in line_items_by_sheet[sheet]:
                errors.append(f"'{item}' is not a valid line item in {sheet}."); continue
            if not isinstance(amount, (int, float)) or isinstance(amount, bool):
                errors.append(f"Amount for '{item}' is not numeric."); continue
            if op not in cls.ALLOWED_OPS:
                errors.append(f"Invalid operation '{op}' for '{item}'."); continue
            valid.append({"sheet": sheet, "line_item": item, "amount": float(amount), "operation": op})

        if not valid:
            return [], False, "Reject: " + "; ".join(errors) if errors else "Nothing valid to apply."
        return valid, False, ("; ".join(errors) if errors else None)


# ============================================================================
# TRANSACTION LOG SERVICE
# ============================================================================
class TransactionLogService:
    @staticmethod
    def record_applied(wb, fiscal_year, quarter, original_txn, entry):
        ExcelService.append_log(wb, fiscal_year, quarter, original_txn, entry["sheet"],
                                 entry["line_item"], entry["amount"], entry["operation"], "Applied")

    @staticmethod
    def record_rejected(wb, fiscal_year, quarter, original_txn, reason):
        ExcelService.append_log(wb, fiscal_year, quarter, original_txn, "-", "-", 0, "-", f"Rejected: {reason}")


# ============================================================================
# DASHBOARD / REPORT SERVICE
# ============================================================================
class DashboardService:

    @classmethod
    def statement_snapshot(cls, wb, fiscal_year, quarter):
        pl = ExcelService.get_period_values(wb, "Profit & Loss", fiscal_year, quarter)
        bs = ExcelService.get_period_values(wb, "Balance Sheet", fiscal_year, quarter)
        cf = ExcelService.get_period_values(wb, "Cash Flow", fiscal_year, quarter)
        return {
            "Profit & Loss": {**pl, **AccountingService.pl_totals(pl)},
            "Balance Sheet": {**bs, **AccountingService.bs_totals(bs)},
            "Cash Flow": {**cf, **AccountingService.cf_totals(cf)},
        }

    @classmethod
    def dashboard_summary(cls, wb, fiscal_year, quarter):
        periods = ExcelService.list_periods(wb)
        if (fiscal_year, quarter) not in periods:
            return None
        snap = cls.statement_snapshot(wb, fiscal_year, quarter)
        pl = snap["Profit & Loss"]

        idx = QUARTER_ORDER.index(quarter)
        prev_q = QUARTER_ORDER[idx - 1] if idx > 0 else None
        prev_pl = ExcelService.get_period_values(wb, "Profit & Loss", fiscal_year, prev_q) if prev_q else {}
        prev_pl_totals = AccountingService.pl_totals(prev_pl) if prev_pl else {}

        yoy_pl = ExcelService.get_period_values(wb, "Profit & Loss", fiscal_year - 1, quarter)
        yoy_pl_totals = AccountingService.pl_totals(yoy_pl) if yoy_pl else {}

        def growth(cur, prev):
            if not prev:
                return None
            return round((cur - prev) / abs(prev) * 100, 1)

        expense_rows = ["Cost of Materials Consumed", "Purchases of Stock-in-Trade", "Changes in Inventories",
                        "Employee Benefits Expense", "Finance Costs",
                        "Depreciation & Amortisation Expenses", "Other Expenses"]
        expense_breakdown = {r: pl.get(r, 0) for r in expense_rows if pl.get(r, 0)}

        return {
            "fiscal_year": fiscal_year, "quarter": quarter,
            "total_income": pl.get("Total Income", 0),
            "profit_after_tax": pl.get("Profit After Tax", 0),
            "net_margin": pl.get("Net Margin %", 0),
            "total_assets": snap["Balance Sheet"].get("TOTAL ASSETS", 0),
            "closing_cash": snap["Cash Flow"].get("Closing Cash", 0),
            "qoq_income": growth(pl.get("Total Income", 0), prev_pl_totals.get("Total Income")),
            "qoq_pat": growth(pl.get("Profit After Tax", 0), prev_pl_totals.get("Profit After Tax")),
            "yoy_income": growth(pl.get("Total Income", 0), yoy_pl_totals.get("Total Income")),
            "yoy_pat": growth(pl.get("Profit After Tax", 0), yoy_pl_totals.get("Profit After Tax")),
            "expense_breakdown": expense_breakdown,
        }

    @classmethod
    def report(cls, wb, fiscal_year):
        """Q1-Q4 + H1/9M/Annual roll-up of P&L (flow items only) with QoQ/YoY,
        mirroring the original dashboard's trend logic. Balance Sheet / Cash
        Flow are point-in-time, so they are reported per-quarter only (see
        /api/statement) rather than summed into cumulative columns."""
        periods = {q: fy for fy, q in ExcelService.list_periods(wb) if fy == fiscal_year}
        available = [q for q in QUARTER_ORDER if q in periods]
        quarter_vals = {q: ExcelService.get_period_values(wb, "Profit & Loss", fiscal_year, q) for q in available}
        quarter_totals = {q: {**quarter_vals[q], **AccountingService.pl_totals(quarter_vals[q])} for q in available}

        prev_year_vals = {q: ExcelService.get_period_values(wb, "Profit & Loss", fiscal_year - 1, q) for q in QUARTER_ORDER}
        prev_year_totals = {q: {**prev_year_vals[q], **AccountingService.pl_totals(prev_year_vals[q])}
                             for q in QUARTER_ORDER if prev_year_vals[q]}

        def sum_rows(qs, row_names):
            return {name: sum(quarter_totals[q].get(name, 0) for q in qs) for name in row_names}

        all_row_names = set()
        for v in quarter_totals.values():
            all_row_names |= set(v.keys())

        columns, cumulative = [], {}
        for q in QUARTER_ORDER:
            if q in available:
                columns.append(q)
            if q == "Q2" and all(x in available for x in ("Q1", "Q2")):
                cumulative["H1 (6M)"] = sum_rows(["Q1", "Q2"], all_row_names)
                columns.append("H1 (6M)")
            if q == "Q3" and all(x in available for x in ("Q1", "Q2", "Q3")):
                cumulative["9M"] = sum_rows(["Q1", "Q2", "Q3"], all_row_names)
                columns.append("9M")
            if q == "Q4" and all(x in available for x in QUARTER_ORDER):
                cumulative["FY (Annual)"] = sum_rows(QUARTER_ORDER, all_row_names)
                columns.append("FY (Annual)")

        all_period_values = {**quarter_totals, **cumulative}
        line_items = [{"particular": name, "values": {col: all_period_values.get(col, {}).get(name, 0)
                                                       for col in columns}}
                      for name in sorted(all_row_names)]

        growth = {}
        for li in line_items:
            name = li["particular"]
            growth[name] = {}
            for i, q in enumerate(available):
                cur = quarter_totals[q].get(name, 0)
                qoq = None
                if i > 0:
                    prev_val = quarter_totals[available[i - 1]].get(name, 0)
                    if prev_val:
                        qoq = round((cur - prev_val) / abs(prev_val) * 100, 1)
                yoy = None
                prev_year_val = prev_year_totals.get(q, {}).get(name)
                if prev_year_val:
                    yoy = round((cur - prev_year_val) / abs(prev_year_val) * 100, 1)
                growth[name][q] = {"qoq": qoq, "yoy": yoy}

        return {"fiscal_year": fiscal_year, "columns": columns, "line_items": line_items, "growth": growth}


# ============================================================================
# STYLED DOWNLOAD WORKBOOK
# The download is always the live workbook itself - cloned from the master
# template, only values changed - never a newly generated workbook.
# ============================================================================
_HEADER_FILL = PatternFill("solid", fgColor="111827")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def get_raw_download_bytes():
    """The workbook exactly as stored - every sheet in it is a clone of the
    master template with only its values changed, per the spec."""
    _, raw = _get_file_meta()
    if raw is not None:
        return raw
    wb = ExcelService.load_live_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================================
# API
# ============================================================================
app = FastAPI(title="Finance Management System - AI Accounting Assistant")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.exception_handler(Exception)
async def _all_exceptions(request, exc: Exception):
    """Any exception that isn't already an HTTPException (missing env vars,
    template not found, GitHub API errors, etc.) becomes a readable JSON
    error instead of a blank 500 with no message - so the browser console
    and Network tab actually tell you what went wrong."""
    return JSONResponse(status_code=500, content={"detail": f"{type(exc).__name__}: {exc}"})


class LoginIn(BaseModel):
    name: str
    access_code: str


class ChatIn(BaseModel):
    fiscal_year: int
    quarter: str
    message: str
    entered_by: str


class UndoIn(BaseModel):
    fiscal_year: int
    quarter: str


@app.post("/api/login")
def login(data: LoginIn):
    if data.access_code != ACCESS_CODE:
        return {"success": False, "message": "Incorrect access code."}
    if not data.name.strip():
        return {"success": False, "message": "Please enter your name."}
    return {"success": True, "name": data.name.strip()}


@app.get("/api/line-items")
def line_items():
    """The fixed catalog of sheets/line items from the master template - used
    to render read-only reference lists in the UI if needed."""
    return ExcelService.get_line_items()


@app.get("/api/periods")
def periods():
    wb = ExcelService.load_live_workbook()
    return {"periods": [{"fiscal_year": fy, "quarter": q} for fy, q in ExcelService.list_periods(wb)]}


@app.post("/api/chat")
def chat(data: ChatIn):
    if data.quarter not in QUARTER_ORDER:
        raise HTTPException(400, "quarter must be one of Q1, Q2, Q3, Q4")
    if not data.message.strip():
        raise HTTPException(400, "message is empty")

    line_items_catalog = ExcelService.get_line_items()

    try:
        parsed = GeminiService.parse_transaction(data.message, line_items_catalog)
    except Exception as e:
        raise HTTPException(502, f"Gemini request failed: {e}")

    valid_entries, needs_confirmation, message = ValidationService.validate(parsed, line_items_catalog)

    wb = ExcelService.load_live_workbook()

    if needs_confirmation or not valid_entries:
        TransactionLogService.record_rejected(wb, data.fiscal_year, data.quarter, data.message,
                                               message or "Could not classify this transaction.")
        try:
            ExcelService.save(wb, f"Log rejected transaction ({data.entered_by})")
        except Exception:
            pass  # logging the rejection is best-effort; still tell the user why
        return {"success": False, "needs_confirmation": needs_confirmation,
                "message": message or "Could not classify this transaction.", "entries": []}

    applied = []
    for entry in valid_entries:
        new_value = ExcelService.apply_entry(wb, data.fiscal_year, data.quarter,
                                              entry["sheet"], entry["line_item"],
                                              entry["amount"], entry["operation"])
        TransactionLogService.record_applied(wb, data.fiscal_year, data.quarter, data.message, entry)
        applied.append({**entry, "new_balance": new_value})

    try:
        ExcelService.save(wb, f"AI entry: {data.message[:60]} ({data.entered_by})")
    except Exception as e:
        raise HTTPException(500, f"Could not save the workbook to GitHub: {e}")

    summary = DashboardService.dashboard_summary(wb, data.fiscal_year, data.quarter)
    return {"success": True, "message": message or "Applied.", "entries": applied, "summary": summary}


@app.post("/api/undo")
def undo(data: UndoIn):
    wb = ExcelService.load_live_workbook()
    try:
        result = ExcelService.undo_last(wb, data.fiscal_year, data.quarter)
    except ValueError as e:
        raise HTTPException(400, str(e))
    try:
        ExcelService.save(wb, f"Undo last transaction for {data.quarter} FY{data.fiscal_year}")
    except Exception as e:
        raise HTTPException(500, f"Could not save the workbook to GitHub: {e}")
    return {"success": True, "reversed": result}


@app.get("/api/transactions/{fiscal_year}/{quarter}")
def transactions(fiscal_year: int, quarter: str):
    wb = ExcelService.load_live_workbook()
    return {"transactions": ExcelService.get_log_rows(wb, fiscal_year, quarter)}


@app.get("/api/statement/{fiscal_year}/{quarter}")
def statement(fiscal_year: int, quarter: str):
    wb = ExcelService.load_live_workbook()
    return DashboardService.statement_snapshot(wb, fiscal_year, quarter)


@app.get("/api/dashboard-summary")
def dashboard_summary(fiscal_year: Optional[int] = None, quarter: Optional[str] = None):
    wb = ExcelService.load_live_workbook()
    periods_list = ExcelService.list_periods(wb)
    if not periods_list:
        return {"summary": None}
    if fiscal_year is None or quarter is None:
        fiscal_year, quarter = periods_list[-1]
    return {"summary": DashboardService.dashboard_summary(wb, fiscal_year, quarter), "periods": periods_list}


@app.get("/api/years")
def get_years():
    wb = ExcelService.load_live_workbook()
    return {"years": sorted({fy for fy, _ in ExcelService.list_periods(wb)}, reverse=True)}


@app.get("/api/report/{fiscal_year}")
def get_report(fiscal_year: int):
    wb = ExcelService.load_live_workbook()
    return DashboardService.report(wb, fiscal_year)


@app.get("/api/download")
def download_excel():
    try:
        raw = get_raw_download_bytes()
    except Exception as e:
        raise HTTPException(500, f"Could not read the workbook: {e}")
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finance_report.xlsx"},
    )


@app.get("/api/download-url")
def download_url_route():
    return {"url": get_download_url()}


@app.post("/api/admin/backfill-carry-forward")
def backfill_carry_forward_route():
    """ONE-TIME maintenance route: repairs periods created before Balance
    Sheet / Opening Cash carry-forward existed. Safe to call once; do not
    call a second time (see ExcelService.backfill_carry_forward docstring)."""
    wb = ExcelService.load_live_workbook()
    fixed = ExcelService.backfill_carry_forward(wb)
    if not fixed:
        return {"success": True, "message": "Nothing to backfill (0 or 1 periods exist).", "periods_fixed": []}
    try:
        ExcelService.save(wb, "One-time backfill: carry forward Balance Sheet / Opening Cash")
    except Exception as e:
        raise HTTPException(500, f"Could not save the workbook to GitHub: {e}")
    return {"success": True, "periods_fixed": fixed}
