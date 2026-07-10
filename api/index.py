"""
index.py
--------
Single-file FastAPI app for Vercel's Python serverless function.

This combines what used to be three files (github_helper.py, finance_logic.py,
index.py) into one, because Vercel's Python bundler does not reliably support
sibling-module imports (e.g. `import github_helper`) inside api/*.py entrypoints.
Keeping everything in one file avoids that import problem entirely.
"""
import os
import io
import base64
from datetime import datetime, timezone
from collections import defaultdict
from typing import Dict

import requests
from openpyxl import Workbook, load_workbook

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel


# ============================================================================
# GITHUB HELPER (was github_helper.py)
# ============================================================================
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "")
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "data/finance_data.xlsx")

API_ROOT = "https://api.github.com"

GH_HEADERS = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json",
}

SHEET_NAME = "Entries"
HEADER_ROW = [
    "fiscal_year",
    "quarter",
    "category",
    "particular",
    "amount",
    "entered_by",
    "updated_at",
]


def _contents_url(path: str) -> str:
    return f"{API_ROOT}/repos/{GITHUB_REPO}/contents/{path}"


def _get_file_meta():
    """Returns (sha, raw_bytes) of the workbook, or (None, None) if it doesn't exist yet."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        raise RuntimeError("GITHUB_TOKEN / GITHUB_REPO are not configured on the server.")
    resp = requests.get(
        _contents_url(EXCEL_PATH),
        headers=GH_HEADERS,
        params={"ref": GITHUB_BRANCH},
        timeout=30,
    )
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    data = resp.json()
    raw = base64.b64decode(data["content"])
    return data["sha"], raw


def _blank_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADER_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_rows():
    """Returns a list of dicts, one per stored line-item entry."""
    sha, raw = _get_file_meta()
    if raw is None:
        return []
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    out = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        out.append(dict(zip(header, r)))
    return out


def save_rows(rows: list, commit_message: str):
    """Overwrites the workbook in the GitHub repo with the given list of row dicts."""
    sha, _ = _get_file_meta()

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADER_ROW)
    for r in rows:
        ws.append([r.get(col) for col in HEADER_ROW])

    buf = io.BytesIO()
    wb.save(buf)
    content_b64 = base64.b64encode(buf.getvalue()).decode()

    payload = {
        "message": commit_message,
        "content": content_b64,
        "branch": GITHUB_BRANCH,
    }
    if sha:
        payload["sha"] = sha

    resp = requests.put(_contents_url(EXCEL_PATH), headers=GH_HEADERS, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()


def get_raw_file_bytes():
    """Used for the 'Download Excel' button - returns the exact bytes of the file in the repo."""
    sha, raw = _get_file_meta()
    if raw is None:
        raw = _blank_workbook_bytes()
    return raw


def download_url():
    """Public raw.githubusercontent.com URL, handy for a direct-link download button."""
    return f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{EXCEL_PATH}"


# ============================================================================
# FINANCE LOGIC (was finance_logic.py)
# ============================================================================
QUARTER_ORDER = ["Q1", "Q2", "Q3", "Q4"]

STANDARD_PARTICULARS = {
    "Revenue": [
        "Revenue from Operations",
        "Other Income",
    ],
    "Expenses": [
        "Cost of Materials / COGS",
        "Employee Benefit Expense",
        "Finance Costs",
        "Depreciation & Amortization",
        "Other Expenses",
    ],
    "Other": [],
}

COMPUTED_ROWS = [
    "Total Revenue",
    "Total Expenses",
    "EBITDA",
    "Profit Before Tax (PBT)",
    "Tax Expense",
    "Net Profit",
    "Net Profit Margin %",
]


def _quarter_key(fy, q):
    return (int(fy), q)


def group_by_period(rows):
    """{ (fiscal_year:int, quarter:str): { particular: amount } }"""
    grouped = defaultdict(dict)
    for r in rows:
        fy, q, particular, amount = r["fiscal_year"], r["quarter"], r["particular"], r["amount"]
        if fy is None or q is None or particular is None:
            continue
        grouped[_quarter_key(fy, q)][particular] = float(amount or 0)
    return grouped


def compute_subtotals(period_values: dict, particular_categories: dict):
    """Given {particular: amount} for one period, compute the derived rows."""
    revenue_items = [p for p in particular_categories.get("Revenue", []) if p in period_values]
    expense_items = [p for p in particular_categories.get("Expenses", []) if p in period_values]

    total_revenue = sum(period_values.get(p, 0) for p in revenue_items)
    depreciation = period_values.get("Depreciation & Amortization", 0)
    finance_costs = period_values.get("Finance Costs", 0)
    tax = period_values.get("Tax Expense", 0)

    total_expenses = sum(period_values.get(p, 0) for p in expense_items)

    ebitda = total_revenue - total_expenses + depreciation + finance_costs
    pbt = total_revenue - total_expenses
    net_profit = pbt - tax
    margin = (net_profit / total_revenue * 100) if total_revenue else 0

    return {
        "Total Revenue": total_revenue,
        "Total Expenses": total_expenses,
        "EBITDA": ebitda,
        "Profit Before Tax (PBT)": pbt,
        "Tax Expense": tax,
        "Net Profit": net_profit,
        "Net Profit Margin %": margin,
    }


def _sum_periods(periods_values: list):
    """Sum a list of {particular: amount} dicts into one (for H1 / 9M / FY roll-ups)."""
    out = defaultdict(float)
    for pv in periods_values:
        for k, v in pv.items():
            out[k] += v or 0
    return dict(out)


def build_report(rows, fiscal_year: int, particular_categories: dict = None):
    particular_categories = particular_categories or STANDARD_PARTICULARS
    grouped = group_by_period(rows)

    quarter_raw = {q: grouped.get(_quarter_key(fiscal_year, q), {}) for q in QUARTER_ORDER}
    prev_quarter_raw = {q: grouped.get(_quarter_key(fiscal_year - 1, q), {}) for q in QUARTER_ORDER}

    available_quarters = [q for q in QUARTER_ORDER if quarter_raw[q]]

    columns = list(available_quarters)
    cumulative_raw = {}
    if all(q in available_quarters for q in ["Q1", "Q2"]):
        cumulative_raw["H1 (6M)"] = _sum_periods([quarter_raw["Q1"], quarter_raw["Q2"]])
        columns.append("H1 (6M)")
    if all(q in available_quarters for q in ["Q1", "Q2", "Q3"]):
        cumulative_raw["9M"] = _sum_periods([quarter_raw["Q1"], quarter_raw["Q2"], quarter_raw["Q3"]])
        columns.append("9M")
    if all(q in available_quarters for q in QUARTER_ORDER):
        cumulative_raw["FY (Annual)"] = _sum_periods([quarter_raw[q] for q in QUARTER_ORDER])
        columns.append("FY (Annual)")

    all_period_values = {**{q: quarter_raw[q] for q in available_quarters}, **cumulative_raw}

    all_particulars = []
    for cat in ["Revenue", "Expenses", "Other"]:
        for p in particular_categories.get(cat, []):
            all_particulars.append((cat, p))
    seen = {p for _, p in all_particulars}
    for q in available_quarters:
        for p in quarter_raw[q]:
            if p not in seen:
                all_particulars.append(("Other", p))
                seen.add(p)

    line_items = []
    for cat, particular in all_particulars:
        if particular in COMPUTED_ROWS:
            continue
        values = {col: all_period_values.get(col, {}).get(particular, 0) for col in columns}
        line_items.append({"particular": particular, "category": cat, "values": values})

    computed = {}
    for col in columns:
        computed[col] = compute_subtotals(all_period_values[col], particular_categories)
    for name in COMPUTED_ROWS:
        line_items.append({
            "particular": name,
            "category": "Computed",
            "values": {col: computed[col][name] for col in columns},
        })

    growth = defaultdict(dict)
    quarter_full_values = {q: {**quarter_raw[q], **computed.get(q, {})} for q in available_quarters}
    prev_quarter_full_values = {}
    for q in QUARTER_ORDER:
        base = prev_quarter_raw[q]
        prev_quarter_full_values[q] = {**base, **compute_subtotals(base, particular_categories)} if base else {}

    row_names = [li["particular"] for li in line_items]
    for name in row_names:
        for i, q in enumerate(available_quarters):
            cur = quarter_full_values[q].get(name, 0)
            qoq = None
            if i > 0:
                prev_q = available_quarters[i - 1]
                prev_val = quarter_full_values[prev_q].get(name, 0)
                if prev_val:
                    qoq = (cur - prev_val) / abs(prev_val) * 100
            yoy = None
            prev_year_val = prev_quarter_full_values.get(q, {}).get(name)
            if prev_year_val:
                yoy = (cur - prev_year_val) / abs(prev_year_val) * 100
            growth[name][q] = {"qoq": qoq, "yoy": yoy}

    return {
        "fiscal_year": fiscal_year,
        "columns": columns,
        "line_items": line_items,
        "growth": growth,
    }


def list_available_fiscal_years(rows):
    years = sorted({int(r["fiscal_year"]) for r in rows if r.get("fiscal_year") is not None}, reverse=True)
    return years


def latest_quarter_summary(rows):
    """For the dashboard cards: latest quarter's Total Revenue / Net Profit / QoQ / YoY."""
    grouped = group_by_period(rows)
    if not grouped:
        return None
    latest_key = max(grouped.keys(), key=lambda k: (k[0], QUARTER_ORDER.index(k[1])))
    fy, q = latest_key
    report = build_report(rows, fy)
    rev_row = next((li for li in report["line_items"] if li["particular"] == "Total Revenue"), None)
    np_row = next((li for li in report["line_items"] if li["particular"] == "Net Profit"), None)
    margin_row = next((li for li in report["line_items"] if li["particular"] == "Net Profit Margin %"), None)
    return {
        "fiscal_year": fy,
        "quarter": q,
        "total_revenue": rev_row["values"].get(q, 0) if rev_row else 0,
        "net_profit": np_row["values"].get(q, 0) if np_row else 0,
        "net_margin": margin_row["values"].get(q, 0) if margin_row else 0,
        "qoq_revenue": report["growth"].get("Total Revenue", {}).get(q, {}).get("qoq"),
        "yoy_revenue": report["growth"].get("Total Revenue", {}).get(q, {}).get("yoy"),
        "qoq_net_profit": report["growth"].get("Net Profit", {}).get(q, {}).get("qoq"),
        "yoy_net_profit": report["growth"].get("Net Profit", {}).get(q, {}).get("yoy"),
    }


# ============================================================================
# FASTAPI APP (was index.py)
# ============================================================================
app = FastAPI(title="Finance Management System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

ACCESS_CODE = os.environ.get("ACCESS_CODE", "finance2026")


class LoginIn(BaseModel):
    name: str
    access_code: str


class EntryIn(BaseModel):
    fiscal_year: int
    quarter: str
    entered_by: str
    particulars: Dict[str, Dict[str, float]]


@app.post("/api/login")
def login(data: LoginIn):
    if data.access_code != ACCESS_CODE:
        return {"success": False, "message": "Incorrect access code."}
    if not data.name.strip():
        return {"success": False, "message": "Please enter your name."}
    return {"success": True, "name": data.name.strip()}


@app.get("/api/particulars")
def get_particulars():
    return STANDARD_PARTICULARS


@app.post("/api/entry")
def save_entry(data: EntryIn):
    if data.quarter not in QUARTER_ORDER:
        raise HTTPException(400, "quarter must be one of Q1, Q2, Q3, Q4")

    try:
        rows = load_rows()
    except Exception as e:
        raise HTTPException(500, f"Could not read the workbook from GitHub: {e}")

    rows = [r for r in rows if not (int(r["fiscal_year"]) == data.fiscal_year and r["quarter"] == data.quarter)]

    now = datetime.now(timezone.utc).isoformat()
    for category, items in data.particulars.items():
        for particular, amount in items.items():
            if not particular:
                continue
            rows.append({
                "fiscal_year": data.fiscal_year,
                "quarter": data.quarter,
                "category": category,
                "particular": particular,
                "amount": float(amount or 0),
                "entered_by": data.entered_by,
                "updated_at": now,
            })

    try:
        save_rows(
            rows,
            commit_message=f"Update {data.quarter} FY{data.fiscal_year} financials ({data.entered_by})",
        )
    except Exception as e:
        raise HTTPException(500, f"Could not save the workbook to GitHub: {e}")

    return {"success": True}


@app.get("/api/entry/{fiscal_year}/{quarter}")
def get_entry(fiscal_year: int, quarter: str):
    rows = load_rows()
    filtered = [r for r in rows if int(r["fiscal_year"]) == fiscal_year and r["quarter"] == quarter]
    out = {"Revenue": {}, "Expenses": {}, "Other": {}}
    for r in filtered:
        out.setdefault(r["category"], {})[r["particular"]] = r["amount"]
    return out


@app.get("/api/years")
def get_years():
    rows = load_rows()
    return {"years": list_available_fiscal_years(rows)}


@app.get("/api/report/{fiscal_year}")
def get_report(fiscal_year: int):
    rows = load_rows()
    return build_report(rows, fiscal_year)


@app.get("/api/dashboard-summary")
def dashboard_summary():
    rows = load_rows()
    summary = latest_quarter_summary(rows)
    years = list_available_fiscal_years(rows)
    return {"summary": summary, "years": years, "entry_count": len(rows)}


@app.get("/api/download")
def download_excel():
    try:
        raw = get_raw_file_bytes()
    except Exception as e:
        raise HTTPException(500, f"Could not fetch the workbook from GitHub: {e}")
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=finance_data.xlsx"},
    )


@app.get("/api/download-url")
def download_url_route():
    return {"url": download_url()}
