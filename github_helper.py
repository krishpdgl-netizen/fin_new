"""
github_helper.py
-----------------
All reads/writes of the financial data go through the GitHub Contents API,
so the single source of truth is the .xlsx file sitting in the GitHub repo
(visible in the repo, downloadable, version-controlled by git history).

Required environment variables:
    GITHUB_TOKEN   - a Personal Access Token with "repo" (contents: read/write) scope
    GITHUB_REPO    - "owner/repo-name"
    GITHUB_BRANCH  - branch to commit to (default: "main")
    EXCEL_PATH     - path of the workbook inside the repo (default: "data/finance_data.xlsx")
"""
import os
import base64
import io
import requests
from openpyxl import Workbook, load_workbook

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "")
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "data/finance_data.xlsx")

API_ROOT = "https://api.github.com"

HEADERS = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json",
}

SHEET_NAME = "Entries"
HEADER_ROW = [
    "fiscal_year",   # e.g. 2025
    "quarter",       # Q1 / Q2 / Q3 / Q4
    "category",      # Revenue / Expenses / Other
    "particular",    # line item name, e.g. "Revenue from Operations"
    "amount",        # numeric value
    "entered_by",    # name of the person who entered it
    "updated_at",    # ISO timestamp
]


def _contents_url(path: str) -> str:
    return f"{API_ROOT}/repos/{GITHUB_REPO}/contents/{path}"


def _get_file_meta():
    """Returns (sha, raw_bytes) of the workbook, or (None, None) if it doesn't exist yet."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        raise RuntimeError("GITHUB_TOKEN / GITHUB_REPO are not configured on the server.")
    resp = requests.get(
        _contents_url(EXCEL_PATH),
        headers=HEADERS,
        params={"ref": GITHUB_BRANCH},
        timeout=30,
    )
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    data = resp.json()
    raw = base64.b64decode(data["content"])
    return data["sha"], raw


def _blank_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADER_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_rows():
    """Returns a list of dicts, one per stored line-item entry."""
    sha, raw = _get_file_meta()
    if raw is None:
        return []
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    out = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        out.append(dict(zip(header, r)))
    return out


def save_rows(rows: list, commit_message: str):
    """Overwrites the workbook in the GitHub repo with the given list of row dicts."""
    sha, _ = _get_file_meta()

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADER_ROW)
    for r in rows:
        ws.append([r.get(col) for col in HEADER_ROW])

    buf = io.BytesIO()
    wb.save(buf)
    content_b64 = base64.b64encode(buf.getvalue()).decode()

    payload = {
        "message": commit_message,
        "content": content_b64,
        "branch": GITHUB_BRANCH,
    }
    if sha:
        payload["sha"] = sha

    resp = requests.put(_contents_url(EXCEL_PATH), headers=HEADERS, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()


def get_raw_file_bytes():
    """Used for the 'Download Excel' button - returns the exact bytes of the file in the repo."""
    sha, raw = _get_file_meta()
    if raw is None:
        raw = _blank_workbook_bytes()
    return raw


def download_url():
    """Public raw.githubusercontent.com URL, handy for a direct-link download button."""
    return f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{EXCEL_PATH}"
